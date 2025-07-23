from flask import Flask, render_template, request, redirect, flash
import smtplib
from email.message import EmailMessage
import pandas as pd
import os
from openpyxl import load_workbook
import requests

app = Flask(__name__)
app.secret_key = "Helping-Temp-Key"  # Needed for flash messages

# ----------------------------
# Home Page (News Feed)
# ----------------------------


@app.route("/")
def home():
    api_key = "d7096e55b10749a999cbd46b567594d7"
    url = f"https://newsapi.org/v2/top-headlines?country=us&pageSize=10&apiKey={api_key}"

    try:
        response = requests.get(url)
        all_articles = response.json().get("articles", [])
        articles = [a for a in all_articles if a.get("urlToImage")]
    except:
        articles = []

    return render_template("home.html", articles=articles)


# ----------------------------
# Static Pages
# ----------------------------
@app.route("/about")
def about():
    return render_template("about.html")


@app.route("/services")
def services():
    return render_template("services.html")


# ----------------------------
# Contact Page (Email Handling)
# ----------------------------
@app.route("/contact", methods=["GET", "POST"])
def contact():
    if request.method == "POST":
        name = request.form.get("name")
        email = request.form.get("email")
        subject = request.form.get("subject")
        message = request.form.get("message")

        # Gmail Credentials (App Password must be generated)
        your_email = "helpinghandinsured@gmail.com"
        your_password = "urvdbgpvahpsudpb"  # <-- your App Password, no spaces

        # Email to you
        msg = EmailMessage()
        msg["Subject"] = f"New Inquiry: {subject}"
        msg["From"] = your_email
        msg["To"] = your_email
        msg.set_content(f"Name: {name}\nEmail: {email}\n\nMessage:\n{message}")

        # Auto-reply to customer
        auto_reply = EmailMessage()
        auto_reply["Subject"] = "Thanks for reaching out!"
        auto_reply["From"] = your_email
        auto_reply["To"] = email
        auto_reply.set_content(
            f"Hi {name},\n\nThanks for contacting us. We've received your message and will get back to you soon.\n\nBest,\nHelping Hand Insured"
        )

        try:
            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
                smtp.login(your_email, your_password)
                smtp.send_message(msg)
                smtp.send_message(auto_reply)

            flash("Your message has been sent. We'll get back to you soon!", "success")
        except Exception as e:
            print("❌ Email Send Error:", e)
            flash("Something went wrong. Please try again later.", "danger")

        return redirect("/contact")

    return render_template("contact.html")

# registration page


@app.route("/registration", methods=["GET", "POST"])
def registration():
    if request.method == "POST":
        name = request.form.get("name")
        email = request.form.get("email")
        age_range = request.form.get("age_range")
        passport = request.form.get("passport_country")
        passport_other = request.form.get(
            "other_passport_country") if passport == "Other" else "N/A"
        insurance = request.form.get("insurance_type")
        bundle = request.form.get("bundle")
        car_year = request.form.get("car_year_range")
        medical_type = request.form.get("medical_type")
        vessel_type = request.form.get("vessel_type")
        company = request.form.get("mfo_company_name")

        if insurance == "auto":
            file_name = "Auto Registration Database.xlsx"
            columns = ["Full Name", "Email Address", "Age Range",
                       "Passport Issuing Country", "If Other", "Car Year", "Bundle"]
            values = [name, email, age_range, passport,
                      passport_other, car_year, bundle]
        elif insurance == "medical":
            file_name = "Medical Registration Database.xlsx"
            columns = ["Full Name", "Email Address", "Age Range",
                       "Passport Issuing Country", "If Other", "Coverage Type", "Bundle"]
            values = [name, email, age_range, passport,
                      passport_other, medical_type, bundle]
        elif insurance == "mfo vessel":
            file_name = "MFO Registration Database.xlsx"
            columns = ["Full Name", "Email Adress", "Age Range", "Passport Issuing Country",
                       "If Other", "Vessel Type", "Company Name", "Bundle"]
            values = [name, email, age_range, passport,
                      passport_other, vessel_type, company, bundle]
        else:
            flash("Invalid insurance type.", "danger")
            return redirect("/registration")

        file_path = os.path.join("helpinghand_insured", file_name)

        try:
            if os.path.exists(file_path):
                book = load_workbook(file_path)
                sheet = book.active
                next_row = sheet.max_row + 1
                sheet.cell(row=next_row, column=2).value = next_row - 3
                for idx, val in enumerate(values, start=3):
                    sheet.cell(row=next_row, column=idx).value = val
                book.save(file_path)
            else:
                df = pd.DataFrame([values], columns=columns)
                df.index += 1
                df.to_excel(file_path, startrow=3, startcol=2, index_label="#")
        except Exception as e:
            print("❌ Excel Save Error:", e)
            flash(
                "Error saving registration. Please close the file and try again.", "danger")
            return redirect("/registration")

        # EMAIL SETUP
        your_email = "helpinghandinsured@gmail.com"
        your_password = "urvdbgpvahpsudpb"

        # Email to company
        company_msg = EmailMessage()
        company_msg["Subject"] = f"New Registration - {insurance.title()} Insurance"
        company_msg["From"] = your_email
        company_msg["To"] = your_email
        company_msg.set_content(
            f"New registration submitted:\n\nName: {name}\nEmail: {email}\nInsurance Type: {insurance}\nBundle: {bundle}")

        # Email to user
        user_msg = EmailMessage()
        user_msg["Subject"] = "Registration Received - Helping Hand Insured"
        user_msg["From"] = your_email
        user_msg["To"] = email
        user_msg.set_content(
            f"Hi {name},\n\nThank you for registering for {insurance.title()} Insurance with Helping Hand Insured. We'll be reviewing your request and will contact you soon.\n\nBest regards,\nHelping Hand Insured")

        try:
            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
                smtp.login(your_email, your_password)
                smtp.send_message(company_msg)
                smtp.send_message(user_msg)
        except Exception as e:
            print("❌ Email Send Error:", e)
            flash("Message saved, but confirmation email failed.", "warning")
            return redirect("/registration")

        flash("Registration submitted! A confirmation email has been sent.", "success")
        return redirect("/registration")

    return render_template("registration.html")


# Success Page
@app.route("/success")
def success():
    return '''
    <div style="text-align: center; margin-top: 50px;">
        <h2>Thank you! Your registration request has been received.</h2>
        <br>
        <a href="/registration">
           <button style="padding: 10px 20px; background-color: #0d6efd; color: white; border: none; border-radius: 4px;">
                Back to Registration
           </button>
        </a>
    </div>
    '''


if __name__ == "__main__":
    app.run(debug=True)
